# -*- coding: utf-8 -*-
"""
Gera os COMPONENTES do relatório (Tech Challenge Fase 2 — Qualidade de Vinhos):
  1) Um arquivo Excel (.xlsx) com todas as TABELAS formatadas;
  2) Os GRÁFICOS em PNG (sem pizza/rosca), para embutir no relatório Word.

Todas as estatísticas são calculadas sobre a base LIMPA (sem duplicatas, 1.018 amostras).
Modelagem: Regressão Logística e Random Forest, corte de qualidade >= 6.
Reprodutibilidade: semente fixa = 42.
"""
import os
import warnings

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.model_selection import train_test_split, GridSearchCV, StratifiedKFold
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import (accuracy_score, precision_score, recall_score,
                             f1_score, roc_auc_score, confusion_matrix, roc_curve)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ---------------- Configuração ----------------
RANDOM_STATE = 42
CSV = "WineQT.csv"
OUT = "componentes_relatorio"
GRAF = os.path.join(OUT, "graficos")
os.makedirs(GRAF, exist_ok=True)

sns.set_style("whitegrid")
plt.rcParams.update({"figure.dpi": 120, "savefig.dpi": 200, "font.size": 11,
                     "axes.titleweight": "bold"})

AZUL = "#2E5A88"; VERDE = "#4F8A4F"; VERMELHO = "#B0413E"; CINZA = "#7F7F7F"

NOMES_PT = {
    "fixed acidity": "Acidez fixa", "volatile acidity": "Acidez volátil",
    "citric acid": "Ácido cítrico", "residual sugar": "Açúcar residual",
    "chlorides": "Cloretos", "free sulfur dioxide": "Dióxido enxofre livre",
    "total sulfur dioxide": "Dióxido enxofre total", "density": "Densidade",
    "pH": "pH", "sulphates": "Sulfatos", "alcohol": "Teor alcoólico",
}

# ---------------- Dados ----------------
df = pd.read_csv(CSV)
FEATS = [c for c in df.columns if c not in ["Id", "quality"]]
df = df.drop_duplicates(subset=FEATS, keep="first").reset_index(drop=True)  # base limpa 1018
N = len(df)
print(f"Base limpa: {N} amostras")

# ---------------- Tabelas (DataFrames) ----------------
# T1 — Distribuição da nota
vc = df["quality"].value_counts().sort_index()
T1 = pd.DataFrame({"Nota": vc.index, "Quantidade": vc.values})

# T2 — Comparação de cortes
linhas = []
for c in [5, 6, 7, 8]:
    a = int((df["quality"] >= c).sum()); b = N - a
    linhas.append({"Corte (nota >=)": c, "Alta Qualidade": a, "Baixa/Média": b})
T2 = pd.DataFrame(linhas)

# T3 — Correlações de Pearson com a qualidade
corr = df[FEATS + ["quality"]].corr()["quality"].drop("quality").sort_values(ascending=False)
T3 = pd.DataFrame({"Variável": [NOMES_PT[v] for v in corr.index],
                   "Correlação com a qualidade": corr.values})

# T7 — Assimetria
sk = df[FEATS].skew().sort_values(ascending=False)
T7 = pd.DataFrame({"Variável": [NOMES_PT[v] for v in sk.index], "Assimetria": sk.values})

# ---------------- Modelagem (corte 6 e diagnóstico corte 7) ----------------
def treina_avalia(corte):
    y = (df["quality"] >= corte).astype(int)
    X = df.drop(columns=["Id", "quality"])
    Xtr, Xte, ytr, yte = train_test_split(X, y, test_size=0.20,
                                          random_state=RANDOM_STATE, stratify=y)
    sc = StandardScaler(); Xtr_s = sc.fit_transform(Xtr); Xte_s = sc.transform(Xte)
    cv = StratifiedKFold(5, shuffle=True, random_state=RANDOM_STATE)
    log = GridSearchCV(LogisticRegression(class_weight="balanced", max_iter=2000,
                       random_state=RANDOM_STATE), {"C": [.01, .1, 1, 10]},
                       cv=cv, scoring="f1", n_jobs=-1).fit(Xtr_s, ytr).best_estimator_
    rf = GridSearchCV(RandomForestClassifier(class_weight="balanced",
                      random_state=RANDOM_STATE, n_jobs=-1),
                      {"n_estimators": [200, 400], "max_depth": [None, 6, 10],
                       "min_samples_leaf": [1, 2, 5]},
                      cv=cv, scoring="f1", n_jobs=-1).fit(Xtr_s, ytr).best_estimator_
    out = {}
    for nome, m in [("Regressão Logística", log), ("Random Forest", rf)]:
        yp = m.predict(Xte_s); pr = m.predict_proba(Xte_s)[:, 1]
        out[nome] = {
            "metrics": [accuracy_score(yte, yp), precision_score(yte, yp),
                        recall_score(yte, yp), f1_score(yte, yp), roc_auc_score(yte, pr)],
            "cm": confusion_matrix(yte, yp), "proba": pr, "y_test": yte.values,
            "modelo": m, "features": X.columns.tolist(),
        }
    return out, yte

res6, yte6 = treina_avalia(6)
res7, _ = treina_avalia(7)

# T4 — Métricas corte 6
COLS_MET = ["Modelo", "Acurácia", "Precisão", "Recall", "F1", "ROC-AUC"]
T4 = pd.DataFrame([[nome] + res6[nome]["metrics"] for nome in res6], columns=COLS_MET)

# T6 — Diagnóstico corte 7
T6 = pd.DataFrame([[nome] + res7[nome]["metrics"] for nome in res7], columns=COLS_MET)

# T5 — Matrizes de confusão
def cm_row(nome, cm):
    tn, fp, fn, tp = cm.ravel()
    return {"Modelo": nome, "VN (Baixa→Baixa)": tn, "FP (Baixa→Alta)": fp,
            "FN (Alta→Baixa)": fn, "VP (Alta→Alta)": tp}
T5 = pd.DataFrame([cm_row(n, res6[n]["cm"]) for n in res6])

print("Tabelas calculadas.")

# =====================================================================
# GRÁFICOS (PNG) — nenhum de pizza/rosca
# =====================================================================
def salvar(fig, nome):
    caminho = os.path.join(GRAF, nome)
    fig.savefig(caminho, bbox_inches="tight"); plt.close(fig)
    print("  gráfico:", caminho)

# G1 — Distribuição da nota
fig, ax = plt.subplots(figsize=(7.5, 4.5))
ax.bar(T1["Nota"].astype(str), T1["Quantidade"], color=AZUL, edgecolor="black")
for x, v in zip(T1["Nota"].astype(str), T1["Quantidade"]):
    ax.text(x, v + 5, str(v), ha="center", fontsize=9)
ax.set_xlabel("Nota de qualidade"); ax.set_ylabel("Quantidade de vinhos")
ax.set_title("Distribuição da nota de qualidade (base limpa, 1.018)")
salvar(fig, "g1_distribuicao_qualidade.png")

# G2 — Classes binárias no corte 6
alta = int((df["quality"] >= 6).sum()); baixa = N - alta
fig, ax = plt.subplots(figsize=(6, 4.5))
ax.bar(["Baixa/Média\n(nota < 6)", "Alta\n(nota >= 6)"], [baixa, alta],
       color=[VERMELHO, VERDE], edgecolor="black")
for x, v in zip([0, 1], [baixa, alta]):
    ax.text(x, v + 4, f"{v}\n({v/N*100:.1f}%)", ha="center", fontsize=10)
ax.set_ylabel("Quantidade de vinhos")
ax.set_title("Classes binárias com o corte em 6 (equilíbrio 1,16 : 1)")
salvar(fig, "g2_classes_corte6.png")

# G3 — Correlações com a qualidade
co = corr.sort_values()
cores = [VERMELHO if v < 0 else VERDE for v in co.values]
fig, ax = plt.subplots(figsize=(8, 5.5))
ax.barh([NOMES_PT[i] for i in co.index], co.values, color=cores, edgecolor="black")
ax.axvline(0, color="black", linewidth=0.8)
ax.set_xlabel("Correlação de Pearson com a qualidade")
ax.set_title("Correlação das variáveis com a qualidade")
for i, v in enumerate(co.values):
    ax.text(v + (0.01 if v >= 0 else -0.01), i, f"{v:+.3f}",
            va="center", ha="left" if v >= 0 else "right", fontsize=8)
salvar(fig, "g3_correlacoes.png")

# G4 — Matriz de correlação (heatmap)
fig, ax = plt.subplots(figsize=(9, 7.5))
M = df[FEATS + ["quality"]].copy()
M.columns = [NOMES_PT.get(c, "Qualidade") for c in M.columns]
sns.heatmap(M.corr(), annot=True, fmt=".2f", cmap="coolwarm", center=0,
            vmin=-1, vmax=1, square=False, cbar_kws={"label": "Correlação"},
            annot_kws={"size": 7}, ax=ax)
ax.set_title("Matriz de correlação entre as variáveis")
salvar(fig, "g4_matriz_correlacao.png")

# G5 — Curvas ROC (corte 6)
fig, ax = plt.subplots(figsize=(6.5, 6))
for nome, cor in [("Regressão Logística", AZUL), ("Random Forest", VERDE)]:
    fpr, tpr, _ = roc_curve(res6[nome]["y_test"], res6[nome]["proba"])
    auc = res6[nome]["metrics"][4]
    ax.plot(fpr, tpr, color=cor, linewidth=2, label=f"{nome} (AUC={auc:.3f})")
ax.plot([0, 1], [0, 1], "--", color=CINZA, linewidth=1, label="Aleatório")
ax.set_xlabel("Taxa de falsos positivos"); ax.set_ylabel("Taxa de verdadeiros positivos")
ax.set_title("Curvas ROC — corte em 6"); ax.legend(loc="lower right")
salvar(fig, "g5_roc_corte6.png")

# G6 — Matrizes de confusão (heatmaps lado a lado)
fig, axes = plt.subplots(1, 2, figsize=(11, 4.6))
rotulos = ["Baixa/Média", "Alta"]
for ax, nome in zip(axes, ["Regressão Logística", "Random Forest"]):
    sns.heatmap(res6[nome]["cm"], annot=True, fmt="d", cmap="Blues", cbar=False,
                xticklabels=rotulos, yticklabels=rotulos, annot_kws={"size": 13}, ax=ax)
    ax.set_xlabel("Previsto"); ax.set_ylabel("Real"); ax.set_title(nome)
fig.suptitle("Matrizes de confusão — corte em 6 (conjunto de teste, 204 amostras)",
             fontweight="bold")
salvar(fig, "g6_matriz_confusao.png")

# G7 — Importância das variáveis (Random Forest)
rf6 = res6["Random Forest"]["modelo"]; feats = res6["Random Forest"]["features"]
imp = pd.Series(rf6.feature_importances_, index=feats).sort_values()
fig, ax = plt.subplots(figsize=(8, 5.5))
ax.barh([NOMES_PT[i] for i in imp.index], imp.values, color=AZUL, edgecolor="black")
ax.set_xlabel("Importância relativa"); ax.set_title("Importância das variáveis — Random Forest")
for i, v in enumerate(imp.values):
    ax.text(v + 0.002, i, f"{v:.3f}", va="center", fontsize=8)
salvar(fig, "g7_importancia_rf.png")

# G8 — Coeficientes (Regressão Logística)
log6 = res6["Regressão Logística"]["modelo"]
coef = pd.Series(log6.coef_[0], index=feats).sort_values()
cores = [VERMELHO if v < 0 else VERDE for v in coef.values]
fig, ax = plt.subplots(figsize=(8, 5.5))
ax.barh([NOMES_PT[i] for i in coef.index], coef.values, color=cores, edgecolor="black")
ax.axvline(0, color="black", linewidth=0.8)
ax.set_xlabel("Coeficiente (variáveis padronizadas)")
ax.set_title("Coeficientes — Regressão Logística")
salvar(fig, "g8_coeficientes_logreg.png")

print("Gráficos gerados.")

# =====================================================================
# EXCEL (.xlsx) — tabelas formatadas
# =====================================================================
FONTE = "Arial"
fill_hdr = PatternFill("solid", fgColor="2E5A88")
fill_tit = PatternFill("solid", fgColor="1F3864")
fill_zebra = PatternFill("solid", fgColor="EDF2F8")
fina = Side(style="thin", color="BFBFBF")
borda = Border(left=fina, right=fina, top=fina, bottom=fina)
cen = Alignment(horizontal="center", vertical="center")
esq = Alignment(horizontal="left", vertical="center")

wb = Workbook()

def nova_aba(nome, titulo, df_tab, formatos=None, larguras=None, fonte_dados=None):
    """Escreve uma tabela formatada em uma nova aba. formatos: dict col->numfmt."""
    ws = wb.create_sheet(nome)
    ncol = df_tab.shape[1]
    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(1, 1, titulo)
    t.font = Font(name=FONTE, bold=True, size=13, color="FFFFFF")
    t.fill = fill_tit; t.alignment = cen
    # Cabeçalho
    for j, col in enumerate(df_tab.columns, start=1):
        c = ws.cell(3, j, col)
        c.font = Font(name=FONTE, bold=True, color="FFFFFF")
        c.fill = fill_hdr; c.alignment = cen; c.border = borda
    # Dados
    for i, (_, row) in enumerate(df_tab.iterrows()):
        r = 4 + i
        for j, col in enumerate(df_tab.columns, start=1):
            c = ws.cell(r, j, row[col])
            c.font = Font(name=FONTE)
            c.alignment = esq if j == 1 else cen
            c.border = borda
            if i % 2 == 1:
                c.fill = fill_zebra
            if formatos and col in formatos:
                c.number_format = formatos[col]
    # Larguras
    for j in range(1, ncol + 1):
        L = get_column_letter(j)
        ws.column_dimensions[L].width = (larguras[j-1] if larguras else 18)
    # Fonte dos dados (nota de rodapé)
    if fonte_dados:
        rr = 4 + len(df_tab) + 1
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=ncol)
        f = ws.cell(rr, 1, fonte_dados)
        f.font = Font(name=FONTE, italic=True, size=8, color="7F7F7F")
    ws.sheet_view.showGridLines = False
    return ws

FONTE_MODELO = "Fonte: pipeline Python (scikit-learn), conjunto de teste de 204 amostras, semente fixa = 42."
FONTE_BASE = "Fonte: WineQT (base limpa, 1.018 amostras após remoção de duplicatas)."

# Aba 1 — Distribuição (com % por fórmula e total)
ws = nova_aba("Distribuicao", "Distribuição da nota de qualidade",
              T1.assign(**{"Percentual": 0.0}),
              formatos={"Percentual": "0.0%"}, larguras=[12, 16, 14],
              fonte_dados=FONTE_BASE)
n1 = len(T1)
tot_row = 4 + n1
for i in range(n1):
    ws.cell(4 + i, 3).value = f"=B{4+i}/$B${tot_row}"
ws.cell(tot_row, 1, "Total").font = Font(name=FONTE, bold=True)
ws.cell(tot_row, 2).value = f"=SUM(B4:B{tot_row-1})"; ws.cell(tot_row, 2).font = Font(name=FONTE, bold=True)
ws.cell(tot_row, 2).border = borda; ws.cell(tot_row, 1).border = borda
ws.cell(tot_row, 3).value = f"=B{tot_row}/$B${tot_row}"; ws.cell(tot_row, 3).number_format = "0.0%"
ws.cell(tot_row, 3).font = Font(name=FONTE, bold=True); ws.cell(tot_row, 3).border = borda
ws.cell(tot_row, 1).alignment = esq; ws.cell(tot_row, 2).alignment = cen; ws.cell(tot_row, 3).alignment = cen

# Aba 2 — Comparação de cortes (% e razão por fórmula)
T2x = T2.assign(**{"Alta (%)": 0.0, "Baixa (%)": 0.0, "Razão (Alta:Baixa)": 0.0})
T2x = T2x[["Corte (nota >=)", "Alta Qualidade", "Alta (%)", "Baixa/Média", "Baixa (%)", "Razão (Alta:Baixa)"]]
ws = nova_aba("Cortes", "Comparação de pontos de corte (impacto no balanceamento)", T2x,
              formatos={"Alta (%)": "0.0%", "Baixa (%)": "0.0%", "Razão (Alta:Baixa)": "0.00"},
              larguras=[16, 16, 12, 14, 12, 18], fonte_dados=FONTE_BASE)
for i in range(len(T2x)):
    r = 4 + i
    ws.cell(r, 3).value = f"=B{r}/(B{r}+D{r})"
    ws.cell(r, 5).value = f"=D{r}/(B{r}+D{r})"
    ws.cell(r, 6).value = f"=MAX(B{r},D{r})/MIN(B{r},D{r})"
# destacar a linha do corte 6 (i=1)
for j in range(1, 7):
    ws.cell(5, j).fill = PatternFill("solid", fgColor="D9EAD3")
    ws.cell(5, j).font = Font(name=FONTE, bold=True)

# Aba 3 — Correlações
nova_aba("Correlacoes", "Correlação de Pearson de cada variável com a qualidade", T3,
         formatos={"Correlação com a qualidade": "0.000"}, larguras=[26, 26],
         fonte_dados=FONTE_BASE)

# Aba 4 — Métricas corte 6
nova_aba("Metricas_Corte6", "Desempenho dos modelos — corte em 6", T4,
         formatos={c: "0.000" for c in COLS_MET[1:]}, larguras=[22, 12, 12, 12, 12, 12],
         fonte_dados=FONTE_MODELO)

# Aba 5 — Matrizes de confusão
nova_aba("Matrizes_Confusao", "Matrizes de confusão — corte em 6 (teste, 204 amostras)", T5,
         larguras=[22, 18, 16, 16, 16], fonte_dados=FONTE_MODELO)

# Aba 6 — Diagnóstico corte 7
nova_aba("Diagnostico_Corte7", "Diagnóstico do corte em 7 (justifica a escolha do 6)", T6,
         formatos={c: "0.000" for c in COLS_MET[1:]}, larguras=[22, 12, 12, 12, 12, 12],
         fonte_dados=FONTE_MODELO)

# Aba 7 — Assimetria
nova_aba("Assimetria", "Assimetria (skewness) das variáveis", T7,
         formatos={"Assimetria": "0.00"}, larguras=[26, 16], fonte_dados=FONTE_BASE)

# Capa/índice
cover = wb["Sheet"]; cover.title = "Indice"
cover.sheet_view.showGridLines = False
cover.merge_cells("A1:C1")
c = cover.cell(1, 1, "Tech Challenge Fase 2 — Qualidade de Vinhos")
c.font = Font(name=FONTE, bold=True, size=15, color="FFFFFF"); c.fill = fill_tit; c.alignment = cen
cover.merge_cells("A2:C2")
cover.cell(2, 1, "Tabelas de apoio ao relatório (base limpa, 1.018 amostras | semente 42)").font = Font(name=FONTE, italic=True, color="595959")
indice = [("Distribuicao", "Distribuição da nota de qualidade"),
          ("Cortes", "Comparação de pontos de corte"),
          ("Correlacoes", "Correlações com a qualidade"),
          ("Metricas_Corte6", "Desempenho dos modelos (corte 6)"),
          ("Matrizes_Confusao", "Matrizes de confusão (corte 6)"),
          ("Diagnostico_Corte7", "Diagnóstico do corte 7"),
          ("Assimetria", "Assimetria das variáveis")]
cover.cell(4, 1, "Aba").font = Font(name=FONTE, bold=True, color="FFFFFF")
cover.cell(4, 1).fill = fill_hdr; cover.cell(4, 2).fill = fill_hdr
cover.cell(4, 2, "Conteúdo").font = Font(name=FONTE, bold=True, color="FFFFFF")
for k, (aba, desc) in enumerate(indice):
    cover.cell(5 + k, 1, aba).font = Font(name=FONTE)
    cover.cell(5 + k, 2, desc).font = Font(name=FONTE)
cover.column_dimensions["A"].width = 22; cover.column_dimensions["B"].width = 42
wb.move_sheet("Indice", -(len(wb.sheetnames) - 1))

xlsx_path = os.path.join(OUT, "Tech_Challenge_Fase2_Tabelas.xlsx")
wb.save(xlsx_path)
print("Excel salvo:", xlsx_path)
